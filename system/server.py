from __future__ import annotations

import cgi
import io
import json
import math
import os
import re
import xml.etree.ElementTree as ET
from http.server import SimpleHTTPRequestHandler, ThreadingHTTPServer
from urllib.parse import parse_qs, urlparse

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles.colors import COLOR_INDEX


PORT = int(os.environ.get("PORT", "8000"))
HOST = os.environ.get("HOST", "127.0.0.1")
PREVIEW_ROWS = 40
PREVIEW_COLS = 24

os.chdir(os.path.dirname(os.path.abspath(__file__)))
SETTINGS_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "wrexcel-settings.json")
SETTINGS_RETENTION_DAYS = 30


def parse_number(value: object) -> float:
    if pd.isna(value):
        return math.nan
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    if not text:
        return math.nan
    cleaned = re.sub(r"[^\d.\-]", "", text)
    return float(cleaned) if cleaned else math.nan


def clean_cell(value: object) -> object:
    if pd.isna(value):
        return ""
    if isinstance(value, pd.Timestamp):
        return value.strftime("%Y-%m-%d")
    return value


def read_workbook(file_bytes: bytes) -> dict[str, pd.DataFrame]:
    excel = pd.ExcelFile(io.BytesIO(file_bytes))
    return {
        sheet_name: pd.read_excel(io.BytesIO(file_bytes), sheet_name=sheet_name, header=None)
        for sheet_name in excel.sheet_names
    }


def suggest_row_bounds(df: pd.DataFrame) -> tuple[int, int]:
    non_empty_counts = df.notna().sum(axis=1)
    likely_rows = non_empty_counts[non_empty_counts >= 2]
    if likely_rows.empty:
        return 1, max(1, len(df))
    return int(likely_rows.index.min() + 1), int(likely_rows.index.max() + 1)


def inspect_excel(file_bytes: bytes) -> dict:
    workbook = read_workbook(file_bytes)
    style_workbook = load_workbook(io.BytesIO(file_bytes), data_only=False)
    sheets = []
    for name, df in workbook.items():
        style_sheet = style_workbook[name]
        visible_row_indexes = [
            row_index
            for row_index in range(df.shape[0])
            if not is_hidden_row(style_sheet, row_index + 1)
        ][:PREVIEW_ROWS]
        visible_col_indexes = [
            col_index
            for col_index in range(df.shape[1])
            if not is_hidden_column(style_sheet, col_index)
        ][:PREVIEW_COLS]
        preview = (
            df.iloc[visible_row_indexes, visible_col_indexes].map(clean_cell).values.tolist()
            if visible_row_indexes and visible_col_indexes
            else []
        )
        start_row, end_row = suggest_row_bounds(df)
        sheets.append(
            {
                "name": name,
                "preview": preview,
                "preview_rows": [row_index + 1 for row_index in visible_row_indexes],
                "preview_cols": visible_col_indexes,
                "total_rows": int(df.shape[0]),
                "total_cols": int(df.shape[1]),
                "suggested_start_row": start_row,
                "suggested_end_row": end_row,
            }
        )
    return {"sheets": sheets}


def apply_tint(rgb: str, tint: float) -> str:
    channels = [int(rgb[i : i + 2], 16) for i in (0, 2, 4)]
    adjusted = []
    for channel in channels:
        if tint < 0:
            value = channel * (1 + tint)
        else:
            value = channel * (1 - tint) + 255 * tint
        adjusted.append(max(0, min(255, round(value))))
    return "".join(f"{value:02X}" for value in adjusted)


def theme_palette(workbook) -> list[str]:
    if not workbook.loaded_theme:
        return []
    root = ET.fromstring(workbook.loaded_theme)
    ns = {"a": "http://schemas.openxmlformats.org/drawingml/2006/main"}
    clr_scheme = root.find(".//a:clrScheme", ns)
    if clr_scheme is None:
        return []
    colors = []
    for child in list(clr_scheme):
        srgb = child.find("a:srgbClr", ns)
        sysclr = child.find("a:sysClr", ns)
        if srgb is not None:
            colors.append(srgb.attrib["val"].upper())
        elif sysclr is not None:
            colors.append(sysclr.attrib["lastClr"].upper())
    return colors


def normalize_fill_color(cell, palette: list[str]) -> str | None:
    fill = cell.fill
    if not fill or fill.fill_type is None:
        return None
    color = fill.fgColor
    if color.type == "rgb" and color.rgb:
        return color.rgb[-6:].upper()
    if color.type == "indexed" and color.indexed is not None:
        return COLOR_INDEX[color.indexed][-6:].upper()
    if color.type == "theme" and color.theme is not None:
        if color.theme < len(palette):
            base = palette[color.theme]
            return apply_tint(base, color.tint or 0)
        return None
    return None


def is_hidden_row(sheet, row_number: int) -> bool:
    dimension = sheet.row_dimensions[row_number]
    return bool(dimension.hidden or dimension.height == 0)


def is_hidden_column(sheet, column_index: int) -> bool:
    column_letter = get_column_letter(column_index + 1)
    dimension = sheet.column_dimensions[column_letter]
    return bool(dimension.hidden or dimension.width == 0)


def ensure_visible_column(sheet, column_index: int, label: str) -> None:
    if is_hidden_column(sheet, column_index):
        raise ValueError(f"{label}로 선택한 열이 엑셀에서 숨김 처리되어 있습니다. 숨김 해제 후 다시 선택해주세요.")


def column_colors(
    file_bytes: bytes,
    sheet_name: str,
    column_index: int,
    start_row: int,
    end_row: int,
) -> dict:
    workbook = load_workbook(io.BytesIO(file_bytes), data_only=False)
    if sheet_name not in workbook.sheetnames:
        raise ValueError("선택한 시트를 찾지 못했습니다.")
    sheet = workbook[sheet_name]
    palette = theme_palette(workbook)
    counts: dict[str, int] = {}
    if is_hidden_column(sheet, column_index):
        return {"colors": []}
    for row in range(max(start_row, 1), min(end_row, sheet.max_row) + 1):
        if is_hidden_row(sheet, row):
            continue
        value = sheet.cell(row=row, column=column_index + 1).value
        if value is None:
            continue
        color = normalize_fill_color(sheet.cell(row=row, column=column_index + 1), palette)
        if color:
            counts[color] = counts.get(color, 0) + 1
    return {
        "colors": [
            {"value": color, "label": f"#{color}" if len(color) == 6 else color, "count": count}
            for color, count in counts.items()
        ]
    }


def calculate_excel(
    file_bytes: bytes,
    sheet_name: str,
    product_col: int,
    product_color: str,
    original_col: int,
    original_color: str,
    final_price_col: int,
    final_price_color: str,
    category_col: int | None,
    extra_fields: list[dict],
    reward_fields: list[dict],
    start_row: int,
    end_row: int,
) -> dict:
    workbook = read_workbook(file_bytes)
    if sheet_name not in workbook:
        raise ValueError("선택한 시트를 찾지 못했습니다.")
    df = workbook[sheet_name]
    if any(col >= df.shape[1] for col in (product_col, original_col, final_price_col)):
        raise ValueError("선택한 열이 시트 범위를 벗어났습니다.")

    rows = []
    start_index = max(start_row - 1, 0)
    end_index = min(end_row, len(df))
    style_workbook = load_workbook(io.BytesIO(file_bytes), data_only=False)
    style_sheet = style_workbook[sheet_name]
    palette = theme_palette(style_workbook)

    ensure_visible_column(style_sheet, product_col, "제품명")
    ensure_visible_column(style_sheet, original_col, "기존금액")
    ensure_visible_column(style_sheet, final_price_col, "최종가")
    if category_col is not None:
        ensure_visible_column(style_sheet, category_col, "카테고리")
    for field in extra_fields:
        for key, label in (
            ("source_col", "추가 항목 출력 열"),
            ("operand_col", "추가 항목 뺄 열"),
            ("left_col", "추가 항목 왼쪽 열"),
            ("right_col", "추가 항목 오른쪽 열"),
        ):
            col = field.get(key)
            if isinstance(col, int):
                ensure_visible_column(style_sheet, col, label)
    for field in reward_fields:
        col = field.get("source_col")
        if isinstance(col, int):
            ensure_visible_column(style_sheet, col, "적립금액")

    for row_number, (_, record) in enumerate(df.iloc[start_index:end_index].iterrows(), start=start_row):
        if is_hidden_row(style_sheet, row_number):
            continue
        if product_color:
            current_color = normalize_fill_color(style_sheet.cell(row=row_number, column=product_col + 1), palette)
            if current_color != product_color:
                continue
        if original_color:
            current_color = normalize_fill_color(style_sheet.cell(row=row_number, column=original_col + 1), palette)
            if current_color != original_color:
                continue
        if final_price_color:
            current_color = normalize_fill_color(style_sheet.cell(row=row_number, column=final_price_col + 1), palette)
            if current_color != final_price_color:
                continue
        original_price = parse_number(record.iloc[original_col])
        final_price = parse_number(record.iloc[final_price_col])
        if math.isnan(original_price) or math.isnan(final_price) or original_price <= 0:
            continue
        total_discount_amount = round(original_price - final_price)
        product_name = str(clean_cell(record.iloc[product_col]))
        category_name = str(clean_cell(record.iloc[category_col])) if category_col is not None else ""
        extra_values = {}
        for field in extra_fields:
            field_id = str(field.get("id", ""))
            mode = field.get("mode", "raw")
            if mode == "raw":
                col = field.get("source_col")
                extra_values[field_id] = clean_cell(record.iloc[col]) if isinstance(col, int) and col < df.shape[1] else ""
            elif mode == "original_minus":
                col = field.get("operand_col")
                value = parse_number(record.iloc[col]) if isinstance(col, int) and col < df.shape[1] else math.nan
                extra_values[field_id] = round(original_price - value) if not math.isnan(value) else ""
            elif mode == "column_minus":
                left_col = field.get("left_col")
                right_col = field.get("right_col")
                left = parse_number(record.iloc[left_col]) if isinstance(left_col, int) and left_col < df.shape[1] else math.nan
                right = parse_number(record.iloc[right_col]) if isinstance(right_col, int) and right_col < df.shape[1] else math.nan
                extra_values[field_id] = round(left - right) if not math.isnan(left) and not math.isnan(right) else ""
            else:
                extra_values[field_id] = ""
        reward_values = {}
        total_reward_amount = 0
        for field in reward_fields:
            field_id = str(field.get("id", ""))
            col = field.get("source_col")
            value = parse_number(record.iloc[col]) if isinstance(col, int) and col < df.shape[1] else math.nan
            amount = round(value) if not math.isnan(value) else 0
            reward_values[field_id] = amount
            total_reward_amount += amount
        rows.append(
            {
                "row_id": row_number,
                "product_name": product_name,
                "category_name": category_name,
                "original_price": round(original_price),
                "discount_amount": round(final_price),
                "total_discount_amount": total_discount_amount,
                "discount_rate": round((total_discount_amount / original_price) * 100, 1),
                "extra_values": extra_values,
                "reward_values": reward_values,
                "total_reward_amount": total_reward_amount,
                "effective_price": round(final_price - total_reward_amount),
            }
        )
    if not rows:
        raise ValueError("선택한 범위에서 계산 가능한 숫자 행을 찾지 못했습니다.")
    return {"rows": rows, "extra_fields": extra_fields, "reward_fields": reward_fields}



def read_settings_file() -> dict:
    try:
        with open(SETTINGS_FILE, "r", encoding="utf-8") as file:
            data = json.load(file)
            return data if isinstance(data, dict) else {}
    except Exception:
        return {}


def write_settings_file(data: dict) -> None:
    cutoff = pd.Timestamp.now().timestamp() * 1000 - SETTINGS_RETENTION_DAYS * 24 * 60 * 60 * 1000
    cleaned = {
        key: value
        for key, value in data.items()
        if isinstance(value, dict) and float(value.get("savedAt") or 0) >= cutoff
    }
    with open(SETTINGS_FILE, "w", encoding="utf-8") as file:
        json.dump(cleaned, file, ensure_ascii=False, indent=2)


def save_file_setting(file_key: str, setting: dict) -> None:
    data = read_settings_file()
    setting = dict(setting or {})
    setting["fileKey"] = file_key
    setting["savedAt"] = setting.get("savedAt") or int(pd.Timestamp.now().timestamp() * 1000)
    data[file_key] = setting
    write_settings_file(data)


def load_file_setting(file_key: str) -> dict | None:
    return read_settings_file().get(file_key)


def load_latest_file_setting_by_name(file_name: str) -> dict | None:
    latest = None
    for value in read_settings_file().values():
        if value.get("fileName") != file_name:
            continue
        if latest is None or (value.get("savedAt") or 0) > (latest.get("savedAt") or 0):
            latest = value
    return latest


def delete_file_setting(file_key: str) -> None:
    data = read_settings_file()
    if file_key in data:
        del data[file_key]
        write_settings_file(data)


class DiscountHandler(SimpleHTTPRequestHandler):
    def do_GET(self) -> None:
        parsed = urlparse(self.path)
        query = parse_qs(parsed.query)
        if parsed.path == "/api/settings/load":
            self.respond_json({"settings": load_file_setting(query.get("file_key", [""])[0])}, 200)
            return
        if parsed.path == "/api/settings/latest":
            self.respond_json({"settings": load_latest_file_setting_by_name(query.get("file_name", [""])[0])}, 200)
            return
        super().do_GET()

    def do_POST(self) -> None:
        parsed = urlparse(self.path)
        if parsed.path in {"/api/settings/save", "/api/settings/delete"}:
            try:
                payload = self.parse_json_body()
                file_key = str(payload.get("fileKey") or "")
                if not file_key:
                    raise ValueError("파일 기록 키가 없습니다.")
                if parsed.path == "/api/settings/save":
                    save_file_setting(file_key, payload.get("data") or {})
                    self.respond_json({"ok": True}, 200)
                    return
                delete_file_setting(file_key)
                self.respond_json({"ok": True}, 200)
                return
            except Exception as exc:
                self.respond_json({"error": str(exc)}, 400)
                return
        if parsed.path not in {"/api/inspect", "/api/calculate", "/api/column-colors"}:
            self.send_error(404)
            return
        try:
            fields = self.parse_form()
            uploaded = fields.get("file")
            if not uploaded:
                raise ValueError("업로드된 파일이 없습니다.")
            file_bytes = uploaded[0]
            if parsed.path == "/api/inspect":
                self.respond_json(inspect_excel(file_bytes), 200)
                return
            if parsed.path == "/api/column-colors":
                self.respond_json(
                    column_colors(
                        file_bytes=file_bytes,
                        sheet_name=fields.get("sheet_name", [""])[0],
                        column_index=int(fields.get("column_index", ["-1"])[0]),
                        start_row=int(fields.get("start_row", ["1"])[0]),
                        end_row=int(fields.get("end_row", ["1"])[0]),
                    ),
                    200,
                )
                return
            self.respond_json(
                calculate_excel(
                    file_bytes=file_bytes,
                    sheet_name=fields.get("sheet_name", [""])[0],
                    product_col=int(fields.get("product_col", ["-1"])[0]),
                    product_color=fields.get("product_color", [""])[0],
                    original_col=int(fields.get("original_col", ["-1"])[0]),
                    original_color=fields.get("original_color", [""])[0],
                    final_price_col=int(fields.get("final_price_col", ["-1"])[0]),
                    final_price_color=fields.get("final_price_color", [""])[0],
                    category_col=(
                        int(fields.get("category_col", [""])[0])
                        if fields.get("category_col", [""])[0] != ""
                        else None
                    ),
                    extra_fields=json.loads(fields.get("extra_fields", ["[]"])[0]),
                    reward_fields=json.loads(fields.get("reward_fields", ["[]"])[0]),
                    start_row=int(fields.get("start_row", ["1"])[0]),
                    end_row=int(fields.get("end_row", ["1"])[0]),
                ),
                200,
            )
        except Exception as exc:
            self.respond_json({"error": str(exc)}, 400)

    def parse_json_body(self) -> dict:
        length = int(self.headers.get("content-length", 0))
        if length <= 0:
            return {}
        return json.loads(self.rfile.read(length).decode("utf-8"))

    def parse_form(self) -> dict:
        ctype, pdict = cgi.parse_header(self.headers.get("content-type", ""))
        if ctype != "multipart/form-data":
            raise ValueError("엑셀 파일 업로드 형식이 올바르지 않습니다.")
        pdict["boundary"] = bytes(pdict["boundary"], "utf-8")
        pdict["CONTENT-LENGTH"] = int(self.headers.get("content-length", 0))
        return cgi.parse_multipart(self.rfile, pdict)

    def respond_json(self, payload: dict, status: int) -> None:
        body = json.dumps(payload, ensure_ascii=False).encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)


if __name__ == "__main__":
    server = ThreadingHTTPServer((HOST, PORT), DiscountHandler)
    print(f"Serving on http://{HOST}:{PORT}")
    server.serve_forever()



